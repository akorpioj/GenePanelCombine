import openpyxl
import pandas as pd
import io
from flask import send_file, flash, redirect, url_for, request
from .utils import filter_genes_from_panel_data
from .utils import MAX_PANELS
from .utils import logger

def generate_excel_file(final_unique_gene_set, selected_panel_configs_for_generation, panel_names, panel_full_gene_data, search_term_from_post_form, uploaded_panels=None):
    # Create DataFrame and Excel file
    excel_output = io.BytesIO()
    try:
        with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
            # FIRST: Create the combined filtered gene list as the first sheet
            # Build a mapping from gene symbol to (panel name, list type)
            gene_to_panels = {}
            for config, panel_name, panel_genes in zip(selected_panel_configs_for_generation, panel_names, panel_full_gene_data):
                filtered_genes = filter_genes_from_panel_data(panel_genes, config["list_type"])
                for gene_symbol in filtered_genes:
                    if gene_symbol not in gene_to_panels:
                        gene_to_panels[gene_symbol] = []
                    gene_to_panels[gene_symbol].append((panel_name, config["list_type"]))
            
            # Add user panel file names to gene_to_panels
            if uploaded_panels:
                for sheet_name, gene_list in uploaded_panels:
                    for gene_symbol in gene_list:
                        if gene_symbol not in gene_to_panels:
                            gene_to_panels[gene_symbol] = []
                        # Use the user panel file name as the panel name, and 'User upload' as the list type
                        gene_to_panels[gene_symbol].append((sheet_name, 'User upload'))
            
            combined_rows = []
            for gene_symbol in sorted(list(final_unique_gene_set)):
                # Join all panel names and list types for this gene
                panels = gene_to_panels.get(gene_symbol, [])
                panel_names_str = ", ".join([p[0] for p in panels])
                list_types_str = ", ".join([p[1] for p in panels])
                combined_rows.append({
                    'GeneSymbol': gene_symbol,
                    'Panel(s)': panel_names_str,
                    'List type(s)': list_types_str
                })
            
            df_combined = pd.DataFrame(combined_rows, columns=['GeneSymbol', 'Panel(s)', 'List type(s)'])
            df_combined.to_excel(writer, index=False, sheet_name='Combined list')
            ws_combined = writer.sheets['Combined list']
            
            # Fancy formatting for combined sheet
            from openpyxl.styles import Font, PatternFill, Border, Side
            header_fill = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col_idx, col in enumerate(df_combined.columns, 1):
                cell = ws_combined.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                value = str(df_combined.columns[col_idx-1])
                if not df_combined.empty:
                    first_row_val = str(df_combined.iloc[0, col_idx-1])
                    width = max(len(value), len(first_row_val)) + 2
                else:
                    width = len(value) + 2
                ws_combined.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            for row in ws_combined.iter_rows(min_row=1, max_row=ws_combined.max_row, min_col=1, max_col=ws_combined.max_column):
                for cell in row:
                    cell.border = border
            ws_combined.auto_filter.ref = ws_combined.dimensions

            # SECOND: Write each panel's full gene list to its own sheet
            for idx, (panel_genes, panel_name) in enumerate(zip(panel_full_gene_data, panel_names), 1):
                if panel_genes:
                    # Only keep specified fields
                    keep_fields = [
                        'entity_type', 'entity_name', 'confidence_level', 'publications', 'evidence', 'phenotypes', 'mode_of_inheritance'
                    ]
                    # Clean up values: if value is a list like ["foo"], convert to foo
                    def clean_value(val):
                        if isinstance(val, list) and len(val) == 1:
                            return val[0]
                        if isinstance(val, list):
                            return ', '.join(str(v) for v in val)
                        if isinstance(val, str) and val.startswith("['") and val.endswith("']"):
                            return val[2:-2]
                        return val
                    cleaned = []
                    for gene in panel_genes:
                        row = {k: clean_value(gene.get(k, '')) for k in keep_fields}
                        cleaned.append(row)
                    df_panel = pd.DataFrame(cleaned)
                    # Use a safe sheet name (Excel max 31 chars, no special chars)
                    safe_name = f"Panel {idx}"
                    if panel_name:
                        safe_name = panel_name[:27]  # leave room for idx
                    safe_name = f"{safe_name} ({idx})" if safe_name else f"Panel {idx}"
                    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
                        safe_name = safe_name.replace(ch, '')
                    df_panel.to_excel(writer, index=False, sheet_name=safe_name)
                    # Set column widths and enable autofilter
                    ws = writer.book[safe_name]
                    # Fancy formatting: bold headers, bg color, borders
                    from openpyxl.styles import Font, PatternFill, Border, Side
                    header_fill = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")
                    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    for col_idx, col in enumerate(df_panel.columns, 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        # Set column width
                        value = str(df_panel.columns[col_idx-1])
                        if not df_panel.empty:
                            first_row_val = str(df_panel.iloc[0, col_idx-1])
                            width = max(len(value), len(first_row_val)) + 2
                        else:
                            width = len(value) + 2
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
                    # Borders for all cells
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = border
                    ws.auto_filter.ref = ws.dimensions
            # Add uploaded user panels as separate sheets
            if uploaded_panels:
                for sheet_name, gene_list in uploaded_panels:
                    # Excel sheet names max 31 chars, remove special chars
                    safe_name = sheet_name[:31]
                    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
                        safe_name = safe_name.replace(ch, '')
                    df_user = pd.DataFrame({'Genes': gene_list})
                    df_user.to_excel(writer, index=False, sheet_name=safe_name)
                    ws = writer.book[safe_name]
                    from openpyxl.styles import Font, PatternFill, Border, Side
                    header_fill = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")
                    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    for col_idx, col in enumerate(df_user.columns, 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        value = str(df_user.columns[col_idx-1])
                        if not df_user.empty:
                            first_row_val = str(df_user.iloc[0, col_idx-1])
                            width = max(len(value), len(first_row_val)) + 2
                        else:
                            width = len(value) + 2
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = border
                    ws.auto_filter.ref = ws.dimensions
        
        excel_output.seek(0) # Reset stream position
    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        flash(f"Error generating Excel file: {e}", "error")
        redirect_params = {'search_term': search_term_from_post_form}
        for i in range(1, MAX_PANELS + 1):
            redirect_params[f'selected_panel_id_{i}'] = request.form.get(f'panel_id_{i}')
            redirect_params[f'selected_list_type_{i}'] = request.form.get(f'list_type_{i}')
        return redirect(url_for('main.index', **redirect_params))
   
    return send_file(
        excel_output,
        as_attachment=True,
        download_name='filtered_gene_list.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
