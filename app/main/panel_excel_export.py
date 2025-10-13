"""
Excel Export functionality for Saved Panels
Generates Excel files with multiple sheets:
- Genes sheet with detailed gene information
- Metadata sheet with panel information
- Version History sheet with all versions
"""

import openpyxl
import pandas as pd
import io
from flask import send_file
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.models import SavedPanel, PanelGene, PanelVersion
from app.models import db


def apply_excel_styling(worksheet, df):
    """
    Apply consistent styling to an Excel worksheet
    - Bold headers with background color
    - Borders on all cells
    - Auto-fit column widths
    - Enable autofilter
    """
    # Header styling
    header_fill = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Calculate and set column width
        col_letter = get_column_letter(col_idx)
        max_length = len(str(col_name))
        
        # Check first few rows for content length
        if not df.empty:
            for row_idx in range(min(5, len(df))):
                cell_value = str(df.iloc[row_idx, col_idx - 1])
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
        
        # Set width with limits
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Apply borders to all data cells
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Enable autofilter
    if worksheet.max_row > 0:
        worksheet.auto_filter.ref = worksheet.dimensions
    
    return worksheet


def clean_list_value(value):
    """
    Clean list values from database
    Converts single-item lists to strings, joins multi-item lists
    """
    if value is None:
        return ''
    if isinstance(value, list):
        if len(value) == 0:
            return ''
        elif len(value) == 1:
            return str(value[0])
        else:
            return ', '.join(str(v) for v in value)
    if isinstance(value, str) and value.startswith("['") and value.endswith("']"):
        # Handle string representation of list
        return value[2:-2].replace("', '", ", ")
    return str(value)


def safe_sheet_name(name, max_length=31):
    """
    Create a safe Excel sheet name
    - Max 31 characters
    - Remove special characters that Excel doesn't allow
    """
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    safe_name = str(name)[:max_length]
    for char in invalid_chars:
        safe_name = safe_name.replace(char, '')
    return safe_name or 'Sheet'

def generate_panel_excel(panel_ids, filename=None):
    """
    Generate an Excel file for one or more saved panels
    
    Args:
        panel_ids: List of panel IDs or single panel ID
        filename: Optional filename for the download
        
    Returns:
        Flask send_file response with Excel file
    """
    # Ensure panel_ids is a list
    if not isinstance(panel_ids, list):
        panel_ids = [panel_ids]
    
    # Fetch panels
    panels = SavedPanel.query.filter(SavedPanel.id.in_(panel_ids)).all()
    
    if not panels:
        raise ValueError("No panels found with the provided IDs")
    
    # Generate default filename if not provided
    if not filename:
        if len(panels) == 1:
            filename = f"{panels[0].name}_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            filename = f"panels_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    # Sanitize filename
    for char in ['\\', '/', '*', '?', ':', '[', ']', '<', '>', '|']:
        filename = filename.replace(char, '_')
    
    # Create Excel file in memory
    excel_output = io.BytesIO()
    
    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        
        # ===== SHEET 1: Combined Gene List =====
        if len(panels) > 1:
            combined_genes_data = []
            for panel in panels:
                genes = panel.genes.filter_by(is_active=True).all()
                for gene in genes:
                    combined_genes_data.append({
                        'Gene Symbol': gene.gene_symbol,
                        'Panel Name': panel.name,
                        'Confidence Level': gene.confidence_level or '',
                        'Mode of Inheritance': clean_list_value(gene.mode_of_inheritance),
                        'Phenotype': clean_list_value(gene.phenotype),
                        'Evidence Level': clean_list_value(gene.evidence_level),
                    })
            
            if combined_genes_data:
                df_combined = pd.DataFrame(combined_genes_data)
                df_combined = df_combined.sort_values('Gene Symbol')
                df_combined.to_excel(writer, index=False, sheet_name='Combined Genes')
                ws_combined = writer.sheets['Combined Genes']
                apply_excel_styling(ws_combined, df_combined)
        
        # ===== SHEET 2-N: Individual Panel Gene Lists =====
        for panel_idx, panel in enumerate(panels, 1):
            genes = panel.genes.filter_by(is_active=True).order_by(PanelGene.gene_symbol).all()
            
            if genes:
                genes_data = []
                for gene in genes:
                    genes_data.append({
                        'Gene Symbol': gene.gene_symbol,
                        'Gene Name': gene.gene_name or '',
                        'Ensembl ID': gene.ensembl_id or '',
                        'HGNC ID': gene.hgnc_id or '',
                        'Confidence Level': gene.confidence_level or '',
                        'Mode of Inheritance': clean_list_value(gene.mode_of_inheritance),
                        'Phenotype': clean_list_value(gene.phenotype),
                        'Evidence Level': clean_list_value(gene.evidence_level),
                        'Source Panel ID': gene.source_panel_id or '',
                        'Source List Type': gene.source_list_type or '',
                        'User Notes': gene.user_notes or '',
                        'Custom Confidence': gene.custom_confidence or '',
                        'Modified': 'Yes' if gene.is_modified else 'No',
                        'Added At': gene.added_at.strftime('%Y-%m-%d %H:%M') if gene.added_at else '',
                    })
                
                df_genes = pd.DataFrame(genes_data)
                
                # Create sheet name
                if len(panels) == 1:
                    sheet_name = 'Genes'
                else:
                    sheet_name = safe_sheet_name(f"{panel.name[:25]} Genes")
                
                df_genes.to_excel(writer, index=False, sheet_name=sheet_name)
                ws_genes = writer.sheets[sheet_name]
                apply_excel_styling(ws_genes, df_genes)
        
        # ===== SHEET: Panel Metadata =====
        metadata_data = []
        for panel in panels:
            metadata_data.append({
                'Panel ID': panel.id,
                'Panel Name': panel.name,
                'Description': panel.description or '',
                'Tags': panel.tags or '',
                'Owner': panel.owner.username if panel.owner else '',
                'Status': panel.status.value if panel.status else '',
                'Visibility': panel.visibility.value if panel.visibility else '',
                'Gene Count': panel.gene_count,
                'Source Type': panel.source_type or '',
                'Source Reference': panel.source_reference or '',
                'Current Version': panel.current_version_number,
                'Version Count': panel.version_count,
                'Created At': panel.created_at.strftime('%Y-%m-%d %H:%M') if panel.created_at else '',
                'Updated At': panel.updated_at.strftime('%Y-%m-%d %H:%M') if panel.updated_at else '',
                'Last Accessed': panel.last_accessed_at.strftime('%Y-%m-%d %H:%M') if panel.last_accessed_at else '',
            })
        
        df_metadata = pd.DataFrame(metadata_data)
        df_metadata.to_excel(writer, index=False, sheet_name='Panel Metadata')
        ws_metadata = writer.sheets['Panel Metadata']
        apply_excel_styling(ws_metadata, df_metadata)
        
        # ===== SHEET: Version History =====
        version_data = []
        for panel in panels:
            versions = panel.versions.order_by(PanelVersion.version_number.desc()).all()
            for version in versions:
                version_data.append({
                    'Panel Name': panel.name,
                    'Version Number': version.version_number,
                    'Comment': version.comment or '',
                    'Created By': version.created_by.username if version.created_by else '',
                    'Created At': version.created_at.strftime('%Y-%m-%d %H:%M') if version.created_at else '',
                    'Gene Count': version.gene_count,
                    'Changes Summary': version.changes_summary or '',
                    'Is Protected': 'Yes' if version.is_protected else 'No',
                    'Access Count': version.access_count,
                    'Last Accessed': version.last_accessed_at.strftime('%Y-%m-%d %H:%M') if version.last_accessed_at else '',
                })
        
        if version_data:
            df_versions = pd.DataFrame(version_data)
            df_versions.to_excel(writer, index=False, sheet_name='Version History')
            ws_versions = writer.sheets['Version History']
            apply_excel_styling(ws_versions, df_versions)
    
    # Reset stream position
    excel_output.seek(0)
    
    # Return file for download
    return send_file(
        excel_output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def generate_panel_excel_from_data(panel_data, genes_data, versions_data=None, filename=None):
    """
    Generate an Excel file from panel data dictionaries (useful for API responses)
    
    Args:
        panel_data: Dictionary or list of dictionaries with panel metadata
        genes_data: List of dictionaries with gene information
        versions_data: Optional list of dictionaries with version history
        filename: Optional filename for the download
        
    Returns:
        Flask send_file response with Excel file
    """
    if not filename:
        filename = f"panel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Sanitize filename
    for char in ['\\', '/', '*', '?', ':', '[', ']', '<', '>', '|']:
        filename = filename.replace(char, '_')
    
    excel_output = io.BytesIO()
    
    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        
        # Genes sheet
        if genes_data:
            df_genes = pd.DataFrame(genes_data)
            df_genes.to_excel(writer, index=False, sheet_name='Genes')
            ws_genes = writer.sheets['Genes']
            apply_excel_styling(ws_genes, df_genes)
        
        # Metadata sheet
        if panel_data:
            if not isinstance(panel_data, list):
                panel_data = [panel_data]
            df_metadata = pd.DataFrame(panel_data)
            df_metadata.to_excel(writer, index=False, sheet_name='Panel Metadata')
            ws_metadata = writer.sheets['Panel Metadata']
            apply_excel_styling(ws_metadata, df_metadata)
        
        # Version history sheet
        if versions_data:
            df_versions = pd.DataFrame(versions_data)
            df_versions.to_excel(writer, index=False, sheet_name='Version History')
            ws_versions = writer.sheets['Version History']
            apply_excel_styling(ws_versions, df_versions)
    
    excel_output.seek(0)
    
    return send_file(
        excel_output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def generate_panel_csv(panel_ids, filename=None, include_metadata=True, include_versions=True):
    """
    Generate a CSV file for one or more saved panels
    Combines all data into a single CSV with additional columns
    
    Args:
        panel_ids: List of panel IDs or single panel ID
        filename: Optional filename for the download
        include_metadata: Include panel metadata columns
        include_versions: Include version information
        
    Returns:
        Flask send_file response with CSV file
    """
    # Ensure panel_ids is a list
    if not isinstance(panel_ids, list):
        panel_ids = [panel_ids]
    
    # Fetch panels
    panels = SavedPanel.query.filter(SavedPanel.id.in_(panel_ids)).all()
    
    if not panels:
        raise ValueError("No panels found with the provided IDs")
    
    # Generate default filename if not provided
    if not filename:
        if len(panels) == 1:
            filename = f"{panels[0].name}_export_{datetime.now().strftime('%Y%m%d')}.csv"
        else:
            filename = f"panels_export_{datetime.now().strftime('%Y%m%d')}.csv"
    
    # Sanitize filename
    for char in ['\\', '/', '*', '?', ':', '[', ']', '<', '>', '|']:
        filename = filename.replace(char, '_')
    
    # Collect all gene data
    genes_data = []
    for panel in panels:
        genes = panel.genes.filter_by(is_active=True).order_by(PanelGene.gene_symbol).all()
        for gene in genes:
            row = {
                'Gene Symbol': gene.gene_symbol,
                'Gene Name': gene.gene_name or '',
                'Ensembl ID': gene.ensembl_id or '',
                'HGNC ID': gene.hgnc_id or '',
                'Confidence Level': gene.confidence_level or '',
                'Mode of Inheritance': clean_list_value(gene.mode_of_inheritance),
                'Phenotype': clean_list_value(gene.phenotype),
                'Evidence Level': clean_list_value(gene.evidence_level),
                'Source Panel ID': gene.source_panel_id or '',
                'Source List Type': gene.source_list_type or '',
            }
            
            if include_metadata:
                row.update({
                    'Panel Name': panel.name,
                    'Panel Description': panel.description or '',
                    'Panel Tags': panel.tags or '',
                    'Panel Status': panel.status.value if panel.status else '',
                    'Panel Owner': panel.owner.username if panel.owner else '',
                })
            
            if include_versions:
                row['Current Version'] = panel.current_version_number
                row['Created At'] = panel.created_at.strftime('%Y-%m-%d %H:%M') if panel.created_at else ''
                row['Updated At'] = panel.updated_at.strftime('%Y-%m-%d %H:%M') if panel.updated_at else ''
            
            genes_data.append(row)
    
    # Create DataFrame and CSV
    df = pd.DataFrame(genes_data)
    csv_output = io.StringIO()
    df.to_csv(csv_output, index=False)
    csv_output.seek(0)
    
    # Convert to bytes
    csv_bytes = io.BytesIO(csv_output.getvalue().encode('utf-8'))
    csv_bytes.seek(0)
    
    return send_file(
        csv_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )


def generate_panel_tsv(panel_ids, filename=None, include_metadata=True, include_versions=True):
    """
    Generate a TSV (Tab-Separated Values) file for one or more saved panels
    
    Args:
        panel_ids: List of panel IDs or single panel ID
        filename: Optional filename for the download
        include_metadata: Include panel metadata columns
        include_versions: Include version information
        
    Returns:
        Flask send_file response with TSV file
    """
    # Ensure panel_ids is a list
    if not isinstance(panel_ids, list):
        panel_ids = [panel_ids]
    
    # Fetch panels
    panels = SavedPanel.query.filter(SavedPanel.id.in_(panel_ids)).all()
    
    if not panels:
        raise ValueError("No panels found with the provided IDs")
    
    # Generate default filename if not provided
    if not filename:
        if len(panels) == 1:
            filename = f"{panels[0].name}_export_{datetime.now().strftime('%Y%m%d')}.tsv"
        else:
            filename = f"panels_export_{datetime.now().strftime('%Y%m%d')}.tsv"
    
    # Sanitize filename
    for char in ['\\', '/', '*', '?', ':', '[', ']', '<', '>', '|']:
        filename = filename.replace(char, '_')
    
    # Collect all gene data
    genes_data = []
    for panel in panels:
        genes = panel.genes.filter_by(is_active=True).order_by(PanelGene.gene_symbol).all()
        for gene in genes:
            row = {
                'Gene Symbol': gene.gene_symbol,
                'Gene Name': gene.gene_name or '',
                'Ensembl ID': gene.ensembl_id or '',
                'HGNC ID': gene.hgnc_id or '',
                'Confidence Level': gene.confidence_level or '',
                'Mode of Inheritance': clean_list_value(gene.mode_of_inheritance),
                'Phenotype': clean_list_value(gene.phenotype),
                'Evidence Level': clean_list_value(gene.evidence_level),
                'Source Panel ID': gene.source_panel_id or '',
                'Source List Type': gene.source_list_type or '',
            }
            
            if include_metadata:
                row.update({
                    'Panel Name': panel.name,
                    'Panel Description': panel.description or '',
                    'Panel Tags': panel.tags or '',
                    'Panel Status': panel.status.value if panel.status else '',
                    'Panel Owner': panel.owner.username if panel.owner else '',
                })
            
            if include_versions:
                row['Current Version'] = panel.current_version_number
                row['Created At'] = panel.created_at.strftime('%Y-%m-%d %H:%M') if panel.created_at else ''
                row['Updated At'] = panel.updated_at.strftime('%Y-%m-%d %H:%M') if panel.updated_at else ''
            
            genes_data.append(row)
    
    # Create DataFrame and TSV
    df = pd.DataFrame(genes_data)
    tsv_output = io.StringIO()
    df.to_csv(tsv_output, index=False, sep='\t')
    tsv_output.seek(0)
    
    # Convert to bytes
    tsv_bytes = io.BytesIO(tsv_output.getvalue().encode('utf-8'))
    tsv_bytes.seek(0)
    
    return send_file(
        tsv_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype='text/tab-separated-values'
    )


def generate_panel_json(panel_ids, filename=None, include_genes=True, include_versions=True):
    """
    Generate a JSON file for one or more saved panels
    
    Args:
        panel_ids: List of panel IDs or single panel ID
        filename: Optional filename for the download
        include_genes: Include gene list in export
        include_versions: Include version history
        
    Returns:
        Flask send_file response with JSON file
    """
    import json
    
    # Ensure panel_ids is a list
    if not isinstance(panel_ids, list):
        panel_ids = [panel_ids]
    
    # Fetch panels
    panels = SavedPanel.query.filter(SavedPanel.id.in_(panel_ids)).all()
    
    if not panels:
        raise ValueError("No panels found with the provided IDs")
    
    # Generate default filename if not provided
    if not filename:
        if len(panels) == 1:
            filename = f"{panels[0].name}_export_{datetime.now().strftime('%Y%m%d')}.json"
        else:
            filename = f"panels_export_{datetime.now().strftime('%Y%m%d')}.json"
    
    # Sanitize filename
    for char in ['\\', '/', '*', '?', ':', '[', ']', '<', '>', '|']:
        filename = filename.replace(char, '_')
    
    # Build JSON structure
    export_data = {
        'export_date': datetime.now().isoformat(),
        'export_format': 'json',
        'panel_count': len(panels),
        'panels': []
    }
    
    for panel in panels:
        panel_data = {
            'id': panel.id,
            'name': panel.name,
            'description': panel.description,
            'tags': panel.tags.split(',') if panel.tags else [],
            'status': panel.status.value if panel.status else None,
            'visibility': panel.visibility.value if panel.visibility else None,
            'gene_count': panel.gene_count,
            'source_type': panel.source_type,
            'source_reference': panel.source_reference,
            'owner': {
                'id': panel.owner.id,
                'username': panel.owner.username,
                'full_name': panel.owner.get_full_name()
            } if panel.owner else None,
            'version_count': panel.version_count,
            'current_version': panel.current_version_number,
            'created_at': panel.created_at.isoformat() if panel.created_at else None,
            'updated_at': panel.updated_at.isoformat() if panel.updated_at else None,
        }
        
        # Add genes if requested
        if include_genes:
            genes = panel.genes.filter_by(is_active=True).order_by(PanelGene.gene_symbol).all()
            panel_data['genes'] = []
            for gene in genes:
                panel_data['genes'].append({
                    'gene_symbol': gene.gene_symbol,
                    'gene_name': gene.gene_name,
                    'ensembl_id': gene.ensembl_id,
                    'hgnc_id': gene.hgnc_id,
                    'confidence_level': gene.confidence_level,
                    'mode_of_inheritance': gene.mode_of_inheritance,
                    'phenotype': gene.phenotype,
                    'evidence_level': gene.evidence_level,
                    'source_panel_id': gene.source_panel_id,
                    'source_list_type': gene.source_list_type,
                    'user_notes': gene.user_notes,
                    'custom_confidence': gene.custom_confidence,
                    'is_modified': gene.is_modified,
                    'added_at': gene.added_at.isoformat() if gene.added_at else None,
                })
        
        # Add version history if requested
        if include_versions:
            versions = panel.versions.order_by(PanelVersion.version_number.desc()).all()
            panel_data['versions'] = []
            for version in versions:
                panel_data['versions'].append({
                    'version_number': version.version_number,
                    'comment': version.comment,
                    'gene_count': version.gene_count,
                    'changes_summary': version.changes_summary,
                    'created_by': {
                        'id': version.created_by.id,
                        'username': version.created_by.username
                    } if version.created_by else None,
                    'created_at': version.created_at.isoformat() if version.created_at else None,
                    'is_protected': version.is_protected,
                })
        
        export_data['panels'].append(panel_data)
    
    # Create JSON output
    json_output = io.BytesIO()
    json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
    json_output.write(json_str.encode('utf-8'))
    json_output.seek(0)
    
    return send_file(
        json_output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/json'
    )
