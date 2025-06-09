from flask import render_template, request, send_file, flash, jsonify, redirect, url_for
import requests
import pandas as pd
import io
import logging
import openpyxl
import pytz
from datetime import datetime, timedelta
from app.config_settings import Config
from app.extensions import limiter 
from . import main_bp # Import the Blueprint object defined in __init__.py
from ..models import Visit, PanelDownload, db

# --- Configuration & Constants ---
BASE_API_URL = "https://panelapp.genomicsengland.co.uk/api/v1/"
CONFIDENCE_MAPPING = {
    "Green": 3,
    "Amber": 2,
    "Red": 1,
}
# For displaying labels if needed, though not directly used in this Flask version's filtering logic
# CONFIDENCE_LABELS = {v: k for k, v in CONFIDENCE_MAPPING.items()}
LIST_TYPE_GREEN = "Green genes"
LIST_TYPE_GREEN_AND_AMBER = "Green and Amber genes"
LIST_TYPE_AMBER = "Amber genes"
LIST_TYPE_RED = "Red genes"
LIST_TYPE_ALL = "Whole gene panel"
list_type_options = [LIST_TYPE_ALL, LIST_TYPE_GREEN, LIST_TYPE_GREEN_AND_AMBER, LIST_TYPE_AMBER, LIST_TYPE_RED]
GENE_FILTER = { LIST_TYPE_ALL: [1,2,3],
                LIST_TYPE_GREEN: [3],
                LIST_TYPE_GREEN_AND_AMBER: [2,3],
                LIST_TYPE_AMBER: [2],
                LIST_TYPE_RED: [1]
                }

MAX_PANELS = 3 # Define how many panel selection groups are on the form

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Cache clear control ---
last_cache_clear_time = None
CACHE_CLEAR_INTERVAL = timedelta(hours=2)

@main_bp.route('/clear_cache', methods=['POST'])
@limiter.limit("2 per hour")  # Limit cache clearing
def clear_cache():
    global last_cache_clear_time
    helsinki = pytz.timezone('Europe/Helsinki')
    now = datetime.now(helsinki)
    if last_cache_clear_time is not None and (now - last_cache_clear_time) < CACHE_CLEAR_INTERVAL:
        wait_time = CACHE_CLEAR_INTERVAL - (now - last_cache_clear_time)
        flash(f"Cache can only be cleared every 2 hours. Please wait {wait_time} longer.", "warning")
    else:
        get_all_panels_from_api.cache = None
        get_all_panels_from_api.cache_time = None
        get_all_panels_from_api.next_refresh = None
        last_cache_clear_time = now
        flash("Panel cache cleared!", "success")
    return redirect(url_for('main.index'))

# --- Helper Functions ((largely unchanged, minor logging adjustments) ---

def get_all_panels_from_api():
    """
    Fetches all available panels from the PanelApp API, handling pagination.
    Returns a list of panel dictionaries, sorted by name.
    Caches results in a simple way for the duration of the app run (for a more robust cache, use Flask-Caching).
    The cache is refreshed every day at 7:00 Helsinki time.
    """
    helsinki = pytz.timezone('Europe/Helsinki')
    now = datetime.now(helsinki)
    # Check if cache exists and is still valid
    cache = getattr(get_all_panels_from_api, "cache", None)
    cache_time = getattr(get_all_panels_from_api, "cache_time", None)
    next_refresh = getattr(get_all_panels_from_api, "next_refresh", None)
    if cache is not None and next_refresh is not None and now < next_refresh:
        logger.info("Returning cached panel list.")
        return cache

    # Fetch from API
    panels = []
    url = f"{BASE_API_URL}panels/"
    page_count = 0
    max_pages = 50  # Safety break for pagination

    logger.info("Fetching panel list from API...")
    while url and page_count < max_pages:
        page_count += 1
        try:
            response = requests.get(url, timeout=20)
            response.raise_for_status()
            data = response.json()
            for panel_data in data.get("results", []):
                panels.append({
                    "id": panel_data.get("id"),
                    "name": panel_data.get("name", "N/A"),
                    "version": panel_data.get("version", "N/A"),
                    "disease_group": panel_data.get("disease_group", ""), # Ensure empty string if None for search
                    "disease_sub_group": panel_data.get("disease_sub_group", ""), # Ensure empty string
                    "description": panel_data.get("description", ""),
                    # display_name is now constructed in the route after filtering                    
                })
            url = data.get("next")
        except requests.exceptions.RequestException as e:
            logger.error(f"API Error (get_all_panels): {e}")
            # Do not flash here, let the route handle UI messages based on context
            #flash(f"Error fetching panels from API: {e}", "error")
            get_all_panels_from_api.cache = [] # Cache empty list on error to prevent re-fetch attempts
            get_all_panels_from_api.cache_time = now
            get_all_panels_from_api.next_refresh = now + timedelta(hours=1)
            return [] # Return empty on error
        except ValueError as e:
            logger.error(f"API Error (get_all_panels - JSON parsing): {e}")
            #flash(f"Error parsing panel data from API: {e}", "error")
            get_all_panels_from_api.cache = []
            get_all_panels_from_api.cache_time = now
            get_all_panels_from_api.next_refresh = now + timedelta(hours=1)
            return []

    if page_count == max_pages and url:
        logger.warning("Reached maximum page limit while fetching panels. List might be incomplete.")
        #flash("Panel list might be incomplete due to API pagination limits.", "warning")

    # Store in cache and set next refresh to next 7:00 Helsinki time
    # Calculate next 7:00
    next_7 = now.replace(hour=7, minute=0, second=0, microsecond=0)
    if now >= next_7:
        next_7 = next_7 + timedelta(days=1)
    get_all_panels_from_api.cache = panels
    get_all_panels_from_api.cache_time = now
    get_all_panels_from_api.next_refresh = next_7
    logger.info(f"Fetched and cached {len(panels)} panels. Next refresh at {next_7}.")
    return panels

get_all_panels_from_api.cache = None
get_all_panels_from_api.cache_time = None
get_all_panels_from_api.next_refresh = None

def get_panel_genes_data_from_api(panel_id):
    """
    Fetches gene details for a specific panel ID from the PanelApp API.
    """
    logger.info("get_panel_gene_from_api")
    if not panel_id: return []
    url = f"{BASE_API_URL}panels/{panel_id}/"
    logger.info(f"Fetching genes for panel ID: {panel_id}")
    try:
        response = requests.get(url, timeout=20)
        response.raise_for_status()
        data = response.json()
        return data.get("genes", [])
    except requests.exceptions.RequestException as e:
        logger.error(f"API Error (get_panel_genes_data for panel {panel_id}): {e}")
        flash(f"Error fetching genes for panel ID {panel_id}: {e}", "error")
        return []
    except ValueError as e:
        logger.error(f"API Error (get_panel_genes_data for panel {panel_id} - JSON parsing): {e}")
        flash(f"Error parsing gene data for panel ID {panel_id}: {e}", "error")
        return []

def filter_genes_from_panel_data(panel_genes_data, list_type_selection):
    """
    Filters genes from a panel's gene data based on the selected list type.
    """
    filtered_gene_symbols = []
    if not panel_genes_data: return filtered_gene_symbols

    for gene_info in panel_genes_data:
        gene_symbol = gene_info.get("entity_name")
        confidence = int(gene_info.get("confidence_level"))

        if not gene_symbol: continue

        if confidence in GENE_FILTER[list_type_selection]:
            filtered_gene_symbols.append(gene_symbol)

    return filtered_gene_symbols

# --- Flask Routes ---

@main_bp.route('/')
@limiter.limit("30 per minute")
def index():
    logger.info("index")
    # No more server-side filtering
    return render_template('index.html', 
                           max_panels=MAX_PANELS,
                           list_type_options=list_type_options)

@main_bp.route("/api/panels")
@limiter.limit("10 per minute")
def api_panels():
    logger.info("api_panels")
    # (re)fetch or cache your master list here
    all_panels_raw = get_all_panels_from_api()

    # inject a display_name for the client
    processed = []
    for p in all_panels_raw:
        p2 = p.copy()
        p2["display_name"] = f"{p['name']} (v{p['version']}, ID: {p['id']})"
        processed.append(p2)
    processed.sort(key=lambda x: x["display_name"])
    return jsonify(processed)

@main_bp.route('/generate', methods=['POST'])
@limiter.limit("30 per hour")  # More strict limit for resource-intensive operation
def generate():
    """
    Handles form submission, processes selected panels, filters genes,
    and returns an Excel file.
    """
    final_unique_gene_set = set()
    selected_panel_configs_for_generation = []
    panel_full_gene_data = []  # Store full gene data for each panel
    panel_names = []           # Store panel names for sheet naming

    # Values from the POST form for generating the list
    search_term_from_post_form = request.form.get('search_term_hidden', '') # Get search term if passed from main form

    for i in range(1, MAX_PANELS + 1):
        panel_id_str = request.form.get(f'panel_id_{i}') # Names from the main form
        list_type = request.form.get(f'list_type_{i}')

        if panel_id_str and panel_id_str != "None" and list_type: # "None" is value for empty selection
            try:
                panel_id = int(panel_id_str)
                selected_panel_configs_for_generation.append({
                    "id": panel_id,
                    "list_type": list_type,
                    "form_index": i # For logging/error messages
                })
            except ValueError:
                flash(f"Invalid panel ID received for panel slot {i}.", "error")
                continue # Skip this invalid entry

    if not selected_panel_configs_for_generation:
        flash("No valid panels selected for gene list generation.", "warning")
        # Redirect back to index, trying to preserve search term and selections
        # This requires GET parameters, so we build them.
        redirect_params = {'search_term': search_term_from_post_form}
        for i in range(1, MAX_PANELS + 1):
            redirect_params[f'selected_panel_id_{i}'] = request.form.get(f'panel_id_{i}')
            redirect_params[f'selected_list_type_{i}'] = request.form.get(f'list_type_{i}')
        return redirect(url_for('main.index', **redirect_params))

    logger.info(f"Processing {len(selected_panel_configs_for_generation)} panel configurations for gene list.")

    for idx, config in enumerate(selected_panel_configs_for_generation, 1):
        raw_genes_for_panel = get_panel_genes_data_from_api(config["id"])
        # Save full panel gene data for Excel
        panel_full_gene_data.append(raw_genes_for_panel)
        # Try to get a panel name for the sheet
        panel_name = f"Panel {idx}"
        try:
            # Try to get the panel name from the API data
            all_panels = get_all_panels_from_api()
            match = next((p for p in all_panels if p["id"] == config["id"]), None)
            if match:
                panel_name = match["name"]
        except Exception:
            pass
        panel_names.append(panel_name)
        # Filtered genes for combined list
        if raw_genes_for_panel:
            filtered_genes = filter_genes_from_panel_data(raw_genes_for_panel, config["list_type"])
            for gene_symbol in filtered_genes:
                final_unique_gene_set.add(gene_symbol)
            logger.info(f"Panel ID {config['id']}: Added {len(filtered_genes)} genes.")
    if not final_unique_gene_set:
        flash("No genes found for the selected criteria.", "info")
        redirect_params = {'search_term': search_term_from_post_form}
        for i in range(1, MAX_PANELS + 1):
            redirect_params[f'selected_panel_id_{i}'] = request.form.get(f'panel_id_{i}')
            redirect_params[f'selected_list_type_{i}'] = request.form.get(f'list_type_{i}')
        return redirect(url_for('main.index', **redirect_params))

    # Log the download
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip and ',' in ip:
        ip = ip.split(',')[0].strip()
        
    download = PanelDownload(
        ip_address=ip,
        download_date=datetime.utcnow(),
        panel_ids=','.join(str(config['id']) for config in selected_panel_configs_for_generation),
        list_types=','.join(config['list_type'] for config in selected_panel_configs_for_generation),
        gene_count=len(final_unique_gene_set)
    )
    
    try:
        db.session.add(download)
        db.session.commit()
    except:
        db.session.rollback()
        logger.error("Failed to log panel download")

    logger.info(f"Total unique genes for Excel: {len(final_unique_gene_set)}")
    
    # Create DataFrame and Excel file
    excel_output = io.BytesIO()
    try:
        with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
            # Write each panel's full gene list to its own sheet
            for idx, (panel_genes, panel_name) in enumerate(zip(panel_full_gene_data, panel_names), 1):
                if panel_genes:
                    # Only keep specified fields
                    keep_fields = [
                        'entity_type', 'entity_name', 'confidence_level', 'publications', 'evidence', 'phenotypes', 'mode_of_inheritance'
                    ]
                    # Clean up values: if value is a list like ["foo"], convert to foo
                    def clean_value(val):
                        if isinstance(val, list) and len(val) == 1:
                            return val[0]
                        if isinstance(val, list):
                            return ', '.join(str(v) for v in val)
                        if isinstance(val, str) and val.startswith("['") and val.endswith("']"):
                            return val[2:-2]
                        return val
                    cleaned = []
                    for gene in panel_genes:
                        row = {k: clean_value(gene.get(k, '')) for k in keep_fields}
                        cleaned.append(row)
                    df_panel = pd.DataFrame(cleaned)
                    # Use a safe sheet name (Excel max 31 chars, no special chars)
                    safe_name = f"Panel {idx}"
                    if panel_name:
                        safe_name = panel_name[:27]  # leave room for idx
                    safe_name = f"{safe_name} ({idx})" if safe_name else f"Panel {idx}"
                    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
                        safe_name = safe_name.replace(ch, '')
                    df_panel.to_excel(writer, index=False, sheet_name=safe_name)
                    # Set column widths and enable autofilter
                    ws = writer.book[safe_name]
                    # Fancy formatting: bold headers, bg color, borders
                    from openpyxl.styles import Font, PatternFill, Border, Side
                    header_fill = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")
                    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    for col_idx, col in enumerate(df_panel.columns, 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        # Set column width
                        value = str(df_panel.columns[col_idx-1])
                        if not df_panel.empty:
                            first_row_val = str(df_panel.iloc[0, col_idx-1])
                            width = max(len(value), len(first_row_val)) + 2
                        else:
                            width = len(value) + 2
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
                    # Borders for all cells
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = border
                    ws.auto_filter.ref = ws.dimensions
            # Write the combined filtered gene list with panel names and list types
            # Build a mapping from gene symbol to (panel name, list type)
            gene_to_panels = {}
            for config, panel_name, panel_genes in zip(selected_panel_configs_for_generation, panel_names, panel_full_gene_data):
                filtered_genes = filter_genes_from_panel_data(panel_genes, config["list_type"])
                for gene_symbol in filtered_genes:
                    if gene_symbol not in gene_to_panels:
                        gene_to_panels[gene_symbol] = []
                    gene_to_panels[gene_symbol].append((panel_name, config["list_type"]))
            combined_rows = []
            for gene_symbol in sorted(list(final_unique_gene_set)):
                # Join all panel names and list types for this gene
                panels = gene_to_panels.get(gene_symbol, [])
                panel_names_str = ", ".join([p[0] for p in panels])
                list_types_str = ", ".join([p[1] for p in panels])
                combined_rows.append({
                    'GeneSymbol': gene_symbol,
                    'Panel(s)': panel_names_str,
                    'List type(s)': list_types_str
                })
            df_combined = pd.DataFrame(combined_rows, columns=['GeneSymbol', 'Panel(s)', 'List type(s)'])
            df_combined.to_excel(writer, index=False, sheet_name='Combined list')
            ws_combined = writer.sheets['Combined list']
            # Fancy formatting for combined sheet
            from openpyxl.styles import Font, PatternFill, Border, Side
            header_fill = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col_idx, col in enumerate(df_combined.columns, 1):
                cell = ws_combined.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                value = str(df_combined.columns[col_idx-1])
                if not df_combined.empty:
                    first_row_val = str(df_combined.iloc[0, col_idx-1])
                    width = max(len(value), len(first_row_val)) + 2
                else:
                    width = len(value) + 2
                ws_combined.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            for row in ws_combined.iter_rows(min_row=1, max_row=ws_combined.max_row, min_col=1, max_col=ws_combined.max_column):
                for cell in row:
                    cell.border = border
            ws_combined.auto_filter.ref = ws_combined.dimensions
        excel_output.seek(0) # Reset stream position
    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        flash(f"Error generating Excel file: {e}", "error")
        redirect_params = {'search_term': search_term_from_post_form}
        for i in range(1, MAX_PANELS + 1):
            redirect_params[f'selected_panel_id_{i}'] = request.form.get(f'panel_id_{i}')
            redirect_params[f'selected_list_type_{i}'] = request.form.get(f'list_type_{i}')
        return redirect(url_for('main.index', **redirect_params))
   
    return send_file(
        excel_output,
        as_attachment=True,
        download_name='filtered_gene_list.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

