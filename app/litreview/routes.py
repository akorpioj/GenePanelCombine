"""
Routes for the LitReview blueprint
"""

import io
import csv
import re
import logging
import click
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from flask import render_template, request, flash, redirect, url_for, jsonify, make_response, send_file
from flask_login import login_required, current_user
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

from . import litreview_bp
from .pubmed_service import pubmed_service
from .genie_service import genie_service
from ..models import (
    db, LiteratureSearch, LiteratureArticle, UserArticleAction,
    AuditActionType, LitReviewSession, LitReviewArticleCategory, SearchResult,
)
from ..audit_service import AuditService

log = logging.getLogger(__name__)

# Retention period: searches and article interactions older than this are eligible
# for automated deletion via the `flask litreview cleanup` CLI command / scheduled job.
_RETENTION_DAYS = 365  # 12 months


@litreview_bp.route('/', methods=['GET'])
@login_required
def index():
    """Literature Review main page"""
    # Get recent searches
    recent_searches = LiteratureSearch.query.filter_by(
        user_id=current_user.id
    ).order_by(LiteratureSearch.created_at.desc()).limit(10).all()
    
    # Log page view
    AuditService.log_view(
        resource_type='page',
        resource_id='litreview_index',
        description='Accessed LitReview page'
    )
    
    return render_template('litreview/index.html', recent_searches=recent_searches)


@litreview_bp.route('/search', methods=['GET', 'POST'])
@login_required
def search():
    """Perform literature search"""
    if request.method == 'POST':
        search_term = request.form.get('search_term', '').strip()
        search_type = request.form.get('search_type', 'gene')
        max_results = int(request.form.get('max_results', 50))
        date_range   = request.form.get('date_range') or None
        article_type = request.form.get('article_type') or None
        pub_status   = request.form.get('pub_status') or None
        language     = request.form.get('language') or None
        
        if not search_term:
            flash('Please enter a search term', 'error')
            return redirect(url_for('litreview.index'))
        
        try:
            # Perform PubMed search
            pmids, total_count = pubmed_service.search_by_gene(
                gene_name=search_term,
                max_results=max_results,
                date_range=date_range,
                article_type=article_type,
                pub_status=pub_status,
                language=language,
            )
            
            # Fetch article details
            articles = pubmed_service.fetch_article_details(pmids)
            
            # Save search to database
            search = pubmed_service.save_search(
                user_id=current_user.id,
                search_term=search_term,
                search_type=search_type,
                results_count=len(articles),
                search_params={
                    'max_results': max_results,
                    'date_range': date_range,
                    'article_type': article_type,
                    'pub_status': pub_status,
                    'language': language,
                }
            )
            
            # Save articles
            pubmed_service.save_articles(articles, search.id)
            
            flash(f'Found {total_count} results for "{search_term}"', 'success')
            return redirect(url_for('litreview.search_results', search_id=search.id))
            
        except Exception as e:
            flash(f'Error performing search: {str(e)}', 'error')
            return redirect(url_for('litreview.index'))
    
    return render_template('litreview/search.html')


@litreview_bp.route('/results/<int:search_id>')
@login_required
def search_results(search_id):
    """Display search results"""
    search = LiteratureSearch.query.get_or_404(search_id)
    
    # Verify user owns search
    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))
    
    # Get search results with articles
    results = search.results.order_by('rank').all()

    # Review session context for button state in results.html
    review_session = LitReviewSession.query.filter_by(
        search_id=search_id,
        user_id=current_user.id,
    ).first()

    review_categorized_count = 0
    review_total = len(results)
    if review_session:
        review_categorized_count = LitReviewArticleCategory.query.filter(
            LitReviewArticleCategory.session_id == review_session.id,
            LitReviewArticleCategory.category > 0,
        ).count()

    # Fetch Genie categorizations for colour-coded article cards (non-fatal)
    genie_categories = {}  # str(pmid) -> category int 1-4; absent means unclassified
    genie_ensembl_id = review_session.ensembl_id if review_session else None
    if not genie_ensembl_id:
        try:
            ids = genie_service.lookup_gene(search.search_query)
            genie_ensembl_id = ids[0] if ids else None
        except Exception:
            pass
    if genie_ensembl_id:
        try:
            pmid_cats = genie_service.get_categorizations(genie_ensembl_id)
            genie_categories = {
                str(item['pmid']): item['category']
                for item in pmid_cats
                if item.get('category', 0) > 0
            }
        except Exception:
            log.debug('Could not fetch Genie categorizations for %s', genie_ensembl_id)

    genie_unclassified_count = sum(
        1 for r in results
        if genie_categories.get(str(r.article.pubmed_id), 0) == 0
    )

    # Log view
    AuditService.log_view(
        resource_type='search_results',
        resource_id=str(search_id),
        description=f'Viewed search results for "{search.search_query}"'
    )

    return render_template(
        'litreview/results.html',
        search=search,
        results=results,
        review_session=review_session,
        review_categorized_count=review_categorized_count,
        review_total=review_total,
        genie_categories=genie_categories,
        genie_unclassified_count=genie_unclassified_count,
    )


def _safe_filename(text):
    """Strip characters that are unsafe in filenames."""
    return re.sub(r'[^\w\-]', '_', text)


@litreview_bp.route('/results/<int:search_id>/export/csv')
@login_required
def export_csv(search_id):
    """Download search results as CSV"""
    search = LiteratureSearch.query.get_or_404(search_id)
    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))

    results = search.results.order_by('rank').all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Rank', 'PMID', 'Title', 'Authors', 'Journal', 'Year', 'DOI', 'Abstract', 'MeSH Terms'])
    for r in results:
        a = r.article
        writer.writerow([
            r.rank,
            a.pubmed_id,
            a.title,
            '; '.join(a.authors or []),
            a.journal or '',
            a.publication_date.year if a.publication_date else '',
            a.doi or '',
            a.abstract or '',
            '; '.join(a.mesh_terms or []),
        ])

    safe_term = _safe_filename(search.search_query)
    filename = f"litreview_{safe_term}_{search.created_at.strftime('%Y%m%d')}.csv"
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    return response


@litreview_bp.route('/results/<int:search_id>/export/excel')
@login_required
def export_excel(search_id):
    """Download search results as Excel spreadsheet"""
    search = LiteratureSearch.query.get_or_404(search_id)
    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))

    results = search.results.order_by('rank').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Search Results'

    headers = ['Rank', 'PMID', 'Title', 'Authors', 'Journal', 'Year', 'DOI', 'Abstract', 'MeSH Terms']
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Column widths
    col_widths = [8, 12, 50, 40, 30, 8, 30, 80, 50]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for r in results:
        a = r.article
        ws.append([
            r.rank,
            a.pubmed_id,
            a.title,
            '; '.join(a.authors or []),
            a.journal or '',
            a.publication_date.year if a.publication_date else '',
            a.doi or '',
            a.abstract or '',
            '; '.join(a.mesh_terms or []),
        ])

    # Wrap text in Title and Abstract columns
    for row in ws.iter_rows(min_row=2):
        for cell in (row[2], row[7]):  # Title, Abstract
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_term = _safe_filename(search.search_query)
    filename = f"litreview_{safe_term}_{search.created_at.strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@litreview_bp.route('/article/<int:article_id>')
@login_required
def article_detail(article_id):
    """Display article details"""
    article = LiteratureArticle.query.get_or_404(article_id)
    
    # Log article view
    action = UserArticleAction.query.filter_by(
        user_id=current_user.id,
        article_id=article.id
    ).first()
    if action:
        action.is_viewed = True
        action.view_count = (action.view_count or 0) + 1
        action.last_viewed_at = datetime.utcnow()
    else:
        action = UserArticleAction(
            user_id=current_user.id,
            article_id=article.id,
            is_viewed=True,
            view_count=1
        )
        db.session.add(action)
    db.session.commit()
    
    AuditService.log_view(
        resource_type='article',
        resource_id=article.pubmed_id,
        description=f'Viewed article: {article.title[:100]}'
    )
    
    return render_template('litreview/article.html', article=article)


@litreview_bp.route('/api/search', methods=['POST'])
@login_required
def api_search():
    """API endpoint for literature search"""
    data = request.get_json()
    
    search_term = data.get('search_term', '').strip()
    search_type = data.get('search_type', 'gene')
    max_results = data.get('max_results', 50)
    
    if not search_term:
        return jsonify({'error': 'Search term required'}), 400
    
    try:
        # Perform search
        pmids, total_count = pubmed_service.search_by_gene(
            gene_name=search_term,
            max_results=max_results
        )
        
        # Fetch articles
        articles = pubmed_service.fetch_article_details(pmids[:max_results])
        
        # Save search
        search = pubmed_service.save_search(
            user_id=current_user.id,
            search_term=search_term,
            search_type=search_type,
            results_count=len(articles)
        )
        
        # Save articles
        saved_articles = pubmed_service.save_articles(articles, search.id)
        
        # Format response
        results = [{
        'pmid': article.pubmed_id,
            'title': article.title,
            'authors': article.authors[:3] if article.authors else [],
            'journal': article.journal,
            'publication_date': article.publication_date.isoformat() if article.publication_date else None,
            'abstract': article.abstract[:300] + '...' if article.abstract and len(article.abstract) > 300 else article.abstract,
            'url': url_for('litreview.article_detail', article_id=article.id, _external=True)
        } for article in saved_articles]
        
        return jsonify({
            'success': True,
            'search_id': search.id,
            'total_count': total_count,
            'results_count': len(results),
            'results': results
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@litreview_bp.route('/search-history')
@login_required
def search_history():
    """Display user's search history"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    searches = LiteratureSearch.query.filter_by(
        user_id=current_user.id
    ).order_by(LiteratureSearch.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('litreview/history.html', searches=searches)


# ── Self-service search history deletion ──────────────────────────────────────

@litreview_bp.route('/search/<int:search_id>/delete', methods=['POST'])
@login_required
def delete_search(search_id):
    """Delete a single search record (and its results via cascade)."""
    search = LiteratureSearch.query.get_or_404(search_id)
    if search.user_id != current_user.id:
        flash('Unauthorized', 'error')
        return redirect(url_for('litreview.search_history'))

    query_text = search.search_query
    db.session.delete(search)
    db.session.commit()

    AuditService.log_action(
        action_type=AuditActionType.USER_DELETE,
        action_description=f'User deleted LitReview search: "{query_text}"',
        resource_type='literature_search',
        resource_id=str(search_id),
    )
    flash(f'Search "{query_text}" deleted.', 'success')
    return redirect(url_for('litreview.search_history'))


@litreview_bp.route('/search-history/clear', methods=['POST'])
@login_required
def clear_search_history():
    """Delete ALL search records for the current user."""
    count = LiteratureSearch.query.filter_by(user_id=current_user.id).count()
    LiteratureSearch.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()

    AuditService.log_action(
        action_type=AuditActionType.USER_DELETE,
        action_description=f'User cleared all LitReview search history ({count} record(s))',
        resource_type='literature_search',
        resource_id='all',
    )
    flash(f'Search history cleared ({count} record(s) removed).', 'success')
    return redirect(url_for('litreview.search_history'))


# ── Retention CLI command ──────────────────────────────────────────────────────

@litreview_bp.cli.command('cleanup')
@click.option('--days', default=_RETENTION_DAYS, show_default=True,
              help='Delete searches and article interactions older than this many days.')
@click.option('--dry-run', is_flag=True, default=False,
              help='Show what would be deleted without committing.')
def cleanup_old_data(days, dry_run):
    """Delete LitReview personal data older than the retention period.

    Run this command periodically (e.g. daily cron):
        flask litreview cleanup
        flask litreview cleanup --days 180
        flask litreview cleanup --dry-run
    """
    cutoff = datetime.utcnow() - timedelta(days=days)

    searches_q = LiteratureSearch.query.filter(LiteratureSearch.created_at < cutoff)
    actions_q = UserArticleAction.query.filter(
        UserArticleAction.updated_at < cutoff,
        UserArticleAction.is_saved == False,  # never delete explicitly saved articles
    )

    search_count = searches_q.count()
    action_count = actions_q.count()

    click.echo(f'Retention cutoff: {cutoff.date()} (>{days} days old)')
    click.echo(f'  LiteratureSearch records to delete : {search_count}')
    click.echo(f'  UserArticleAction records to delete: {action_count}  (unsaved only)')

    if dry_run:
        click.echo('Dry-run — no changes committed.')
        return

    searches_q.delete(synchronize_session=False)
    actions_q.delete(synchronize_session=False)
    db.session.commit()
    click.echo(f'Done. {search_count} searches and {action_count} action records deleted.')


# ── Review feature routes ──────────────────────────────────────────────────────

_ENSEMBL_BASE = 'https://www.ensembl.org/Homo_sapiens/Gene/Summary?g='
_OMIM_BASE = 'https://www.omim.org/entry/'


def _build_candidate(ensembl_id: str) -> dict:
    """Fetch detail + OMIM for one Ensembl ID and return a candidate dict."""
    try:
        detail = genie_service.get_gene_detail(ensembl_id)
        omim_id = genie_service.get_omim_id(ensembl_id)
    except Exception:
        detail = None
        omim_id = None

    candidate = {
        'ensembl_id': ensembl_id,
        'display_name': None,
        'chromosome': None,
        'description': None,
        'ensembl_url': f'{_ENSEMBL_BASE}{ensembl_id}',
        'omim_url': f'{_OMIM_BASE}{omim_id}' if omim_id else None,
    }
    if detail:
        candidate['display_name'] = detail.get('display_name')
        candidate['chromosome'] = detail.get('seq_region_name')
        candidate['description'] = detail.get('description')
    return candidate


@litreview_bp.route('/api/gene-lookup')
@login_required
def api_gene_lookup():
    """
    GET /litreview/api/gene-lookup?q=<symbol>

    Calls the Genie API to resolve a gene symbol to one or more Ensembl IDs,
    then fetches detail and OMIM for each candidate in parallel.

    Returns:
        200  {"candidates": [...]}
        200  {"candidates": [], "not_found": true}   — Genie returned nothing
        400  {"error": "..."}                        — missing query param
        500  {"error": "..."}                        — service error
    """
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'error': 'Query parameter "q" is required'}), 400

    try:
        ensembl_ids = genie_service.lookup_gene(q)
    except Exception as exc:
        log.exception('Gene lookup failed for %r', q)
        return jsonify({'error': str(exc)}), 500

    if not ensembl_ids:
        return jsonify({'candidates': [], 'not_found': True})

    # Fetch detail + OMIM for each candidate in parallel
    candidates = [None] * len(ensembl_ids)
    with ThreadPoolExecutor(max_workers=min(len(ensembl_ids), 5)) as pool:
        futures = {
            pool.submit(_build_candidate, eid): idx
            for idx, eid in enumerate(ensembl_ids)
        }
        for future in as_completed(futures):
            candidates[futures[future]] = future.result()

    return jsonify({'candidates': candidates})


@litreview_bp.route('/results/<int:search_id>/review/start')
@login_required
def review_start(search_id):
    """
    GET /litreview/results/<search_id>/review/start

    Gene-confirmation page. If a review session already exists for this
    search + user, redirect straight to the review UI.
    """
    search = LiteratureSearch.query.get_or_404(search_id)

    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))

    # If a session already exists, skip straight to the review page
    existing = LitReviewSession.query.filter_by(
        search_id=search_id,
        user_id=current_user.id,
    ).first()
    if existing:
        return redirect(url_for('litreview.review', search_id=search_id))

    return render_template(
        'litreview/review_start.html',
        search=search,
        search_term=search.search_query,
    )


@litreview_bp.route('/results/<int:search_id>/review/confirm-gene', methods=['POST'])
@login_required
def review_confirm_gene(search_id):
    """
    POST /litreview/results/<search_id>/review/confirm-gene

    Receives ensembl_id (and optional gene_symbol) from the gene-confirmation
    form, creates a LitReviewSession, bulk-inserts LitReviewArticleCategory
    rows for every article in the search (all category=0), then applies any
    previously stored categorizations fetched from the Genie API so that
    an interrupted review can be resumed cleanly.
    """
    search = LiteratureSearch.query.get_or_404(search_id)

    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))

    ensembl_id  = request.form.get('ensembl_id', '').strip()
    gene_symbol = request.form.get('gene_symbol', '').strip() or search.search_query

    if not ensembl_id:
        flash('No Ensembl ID provided. Please confirm a gene before starting the review.', 'error')
        return redirect(url_for('litreview.review_start', search_id=search_id))

    # Prevent duplicate sessions — redirect if one already exists
    existing = LitReviewSession.query.filter_by(
        search_id=search_id,
        user_id=current_user.id,
    ).first()
    if existing:
        return redirect(url_for('litreview.review', search_id=search_id))

    # Create session record
    session_obj = LitReviewSession(
        search_id=search_id,
        user_id=current_user.id,
        ensembl_id=ensembl_id,
        gene_symbol=gene_symbol,
        status='in_progress',
        submitted_to_genie=False,
    )
    db.session.add(session_obj)
    db.session.flush()  # populate session_obj.id

    # Fetch articles ordered by rank
    results = (
        SearchResult.query
        .filter_by(search_id=search_id)
        .order_by(SearchResult.rank)
        .all()
    )

    # Bulk-insert one LitReviewArticleCategory per article (all uncategorized)
    category_rows = {
        r.article_id: LitReviewArticleCategory(
            session_id=session_obj.id,
            article_id=r.article_id,
            pmid=r.article.pubmed_id,
            category=0,
            categorized_at=None,
        )
        for r in results
    }
    db.session.bulk_save_objects(category_rows.values())
    db.session.flush()

    # Pre-populate from Genie API to restore any prior categorizations
    try:
        prior = genie_service.get_categorizations(ensembl_id)
        if prior:
            # Build pmid → article_id reverse map
            pmid_to_article_id = {r.article.pubmed_id: r.article_id for r in results}
            for entry in prior:
                pmid = str(entry.get('pmid', ''))
                cat  = entry.get('category', 0)
                art_id = pmid_to_article_id.get(pmid)
                if art_id and art_id in category_rows:
                    row = category_rows[art_id]
                    if isinstance(cat, int) and 1 <= cat <= 4:
                        row.category = cat
            # Re-save updated rows
            db.session.bulk_save_objects(category_rows.values())
    except Exception:
        log.exception(
            'Could not fetch prior Genie categorizations for %s (search %d); continuing.',
            ensembl_id, search_id
        )
        # Non-fatal — the session is still created, just without prior data

    db.session.commit()

    AuditService.log_action(
        action_type=AuditActionType.SEARCH,
        action_description=(
            f'Started LitReview session for gene {gene_symbol} '
            f'(Ensembl {ensembl_id}) on search {search_id}'
        ),
        resource_type='litreview_session',
        resource_id=str(session_obj.id),
    )

    return redirect(url_for('litreview.review', search_id=search_id))


@litreview_bp.route('/results/<int:search_id>/review')
@login_required
def review(search_id):
    """
    GET /litreview/results/<search_id>/review

    Main categorization page. Redirects to /review/start when no session
    exists yet. Passes articles ordered by rank together with their current
    categories so the template / JS can jump straight to the first uncategorized
    article.
    """
    search = LiteratureSearch.query.get_or_404(search_id)

    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))

    session_obj = LitReviewSession.query.filter_by(
        search_id=search_id,
        user_id=current_user.id,
    ).first()

    if not session_obj:
        return redirect(url_for('litreview.review_start', search_id=search_id))

    # Ordered list of (SearchResult, LitReviewArticleCategory, LiteratureArticle)
    rows = (
        db.session.query(SearchResult, LitReviewArticleCategory, LiteratureArticle)
        .join(LitReviewArticleCategory, LitReviewArticleCategory.article_id == SearchResult.article_id)
        .join(LiteratureArticle, LiteratureArticle.id == SearchResult.article_id)
        .filter(
            SearchResult.search_id == search_id,
            LitReviewArticleCategory.session_id == session_obj.id,
        )
        .order_by(SearchResult.rank)
        .all()
    )

    articles_with_categories = [
        {
            'article': {
                'id':               article.id,
                'pubmed_id':        article.pubmed_id,
                'title':            article.title,
                'abstract':         article.abstract,
                'authors':          article.authors or [],
                'journal':          article.journal,
                'publication_date': article.publication_date.isoformat() if article.publication_date else None,
                'doi':              article.doi,
            },
            'category':  cat_row.category,
            'cat_id':    cat_row.id,
            'rank':      sr.rank,
        }
        for sr, cat_row, article in rows
    ]

    total             = len(articles_with_categories)
    categorized_count = sum(1 for a in articles_with_categories if a['category'] > 0)
    remaining_count   = total - categorized_count

    return render_template(
        'litreview/review.html',
        search=search,
        review_session=session_obj,
        articles_with_categories=articles_with_categories,
        total=total,
        categorized_count=categorized_count,
        remaining_count=remaining_count,
    )


@litreview_bp.route('/results/<int:search_id>/review/categorize', methods=['POST'])
@login_required
def review_categorize(search_id):
    """
    POST /litreview/results/<search_id>/review/categorize
    JSON body: {"article_id": <int>, "category": <int 1-4>}

    Updates the LitReviewArticleCategory for the given article and returns
    the number of still-uncategorized articles in this session.
    """
    search = LiteratureSearch.query.get_or_404(search_id)
    if search.user_id != current_user.id and not current_user.is_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json(silent=True) or {}
    article_id = data.get('article_id')
    category   = data.get('category')

    if article_id is None or category is None:
        return jsonify({'error': 'article_id and category are required'}), 400

    if not isinstance(category, int) or not (1 <= category <= 4):
        return jsonify({'error': 'category must be an integer between 1 and 4'}), 400

    session_obj = LitReviewSession.query.filter_by(
        search_id=search_id,
        user_id=current_user.id,
    ).first()
    if not session_obj:
        return jsonify({'error': 'No review session found'}), 404

    cat_row = LitReviewArticleCategory.query.filter_by(
        id=article_id,
        session_id=session_obj.id,
    ).first()
    if not cat_row:
        log.debug('No category row found for cat_id %d in session %d', article_id, session_obj.id)
        return jsonify({'error': 'Article not found in this session'}), 404
    import datetime as _dt
    cat_row.category       = category
    cat_row.categorized_at = _dt.datetime.now()
    db.session.commit()

    remaining = LitReviewArticleCategory.query.filter_by(
        session_id=session_obj.id,
        category=0,
    ).count()

    return jsonify({'ok': True, 'remaining': remaining})


@litreview_bp.route('/results/<int:search_id>/review/finish', methods=['POST'])
@login_required
def review_finish(search_id):
    """
    POST /litreview/results/<search_id>/review/finish

    Submits all categorized articles (category > 0) to the Genie API in bulk,
    marks the session as complete, and redirects back to the search results.
    """
    search = LiteratureSearch.query.get_or_404(search_id)
    if search.user_id != current_user.id and not current_user.is_admin():
        flash('Unauthorized access', 'error')
        return redirect(url_for('litreview.index'))

    session_obj = LitReviewSession.query.filter_by(
        search_id=search_id,
        user_id=current_user.id,
    ).first()
    if not session_obj:
        flash('No review session found.', 'error')
        return redirect(url_for('litreview.search_results', search_id=search_id))

    # Collect all rows with a non-zero category for submission
    categorized = LitReviewArticleCategory.query.filter(
        LitReviewArticleCategory.session_id == session_obj.id,
        LitReviewArticleCategory.category > 0,
    ).all()

    payload = [
        {'pmid': int(row.pmid), 'category': row.category}
        for row in categorized
    ]

    # Mark session complete first so the UI reflects done state
    # regardless of whether the Genie submission succeeds.
    import datetime as _dt
    session_obj.status             = 'complete'
    session_obj.submitted_to_genie = False
    session_obj.updated_at         = _dt.datetime.now()
    db.session.commit()

    AuditService.log_action(
        action_type=AuditActionType.DATA_EXPORT,
        action_description=(
            f'Finished LitReview session {session_obj.id} for gene '
            f'{session_obj.gene_symbol} (Ensembl {session_obj.ensembl_id}); '
            f'{len(payload)} article(s) queued for Genie submission.'
        ),
        resource_type='litreview_session',
        resource_id=str(session_obj.id),
    )

    if payload:
        try:
            result = genie_service.save_categorizations_bulk(session_obj.ensembl_id, payload)
            added   = len(result.get('added',   []))
            skipped = len(result.get('skipped', []))
            log.info(
                'Genie bulk submit for session %d: %d added, %d skipped',
                session_obj.id, added, skipped,
            )
            session_obj.submitted_to_genie = True
            db.session.commit()
        except Exception:
            log.exception(
                'Genie bulk submit failed for session %d (ensembl %s)',
                session_obj.id, session_obj.ensembl_id,
            )
            flash(
                'Review marked complete, but could not submit to Genie API. '
                'Please contact support to retry the submission.',
                'warning',
            )

    flash(
        f'Review complete! {len(payload)} article(s) submitted to Genie.',
        'success',
    )
    return redirect(url_for('litreview.search_results', search_id=search_id))

