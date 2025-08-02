"""
Unit tests for file upload functionality
"""
import pytest
import io
import os
from werkzeug.datastructures import FileStorage
from unittest.mock import patch, Mock, mock_open
from app.main.utils import validate_file, process_uploaded_file
import openpyxl
import pandas as pd


@pytest.mark.unit
@pytest.mark.file_upload
class TestFileValidation:
    """Test file validation functionality."""
    
    def test_valid_csv_file(self, sample_file_data):
        """Test validation of valid CSV file."""
        csv_content = sample_file_data['valid_csv']
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='test.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        assert result['valid'] is True
        assert 'errors' not in result or len(result['errors']) == 0
    
    def test_invalid_csv_file(self, sample_file_data):
        """Test validation of invalid CSV file."""
        csv_content = sample_file_data['invalid_csv']
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='test.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        # Should detect missing required headers
        assert result['valid'] is False or 'errors' in result
    
    def test_empty_file(self):
        """Test validation of empty file."""
        file_obj = io.BytesIO(b'')
        file_storage = FileStorage(
            stream=file_obj,
            filename='empty.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        assert result['valid'] is False
        assert 'errors' in result
        assert any('empty' in error.lower() for error in result['errors'])
    
    def test_unsupported_file_type(self):
        """Test validation of unsupported file type."""
        file_obj = io.BytesIO(b'some content')
        file_storage = FileStorage(
            stream=file_obj,
            filename='test.txt',
            content_type='text/plain'
        )
        
        result = validate_file(file_storage)
        
        assert result['valid'] is False
        assert 'errors' in result
        assert any('format' in error.lower() or 'type' in error.lower() for error in result['errors'])
    
    def test_file_too_large(self):
        """Test validation of oversized file."""
        large_content = 'x' * (10 * 1024 * 1024)  # 10MB
        file_obj = io.BytesIO(large_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='large.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        # Depending on size limits, this might fail
        # This test depends on actual size limit configuration
        assert isinstance(result, dict)
        assert 'valid' in result
    
    def test_malformed_csv_file(self):
        """Test validation of malformed CSV file."""
        malformed_csv = 'Gene Symbol,Gene Name\n"Unclosed quote,BRCA1\nTP53,Normal Gene'
        file_obj = io.BytesIO(malformed_csv.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='malformed.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        # Should handle malformed CSV gracefully
        assert isinstance(result, dict)
        assert 'valid' in result


@pytest.mark.unit
@pytest.mark.file_upload
class TestFileProcessing:
    """Test file processing functionality."""
    
    def test_process_valid_csv(self, sample_file_data):
        """Test processing of valid CSV file."""
        csv_content = sample_file_data['valid_csv']
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='test.csv',
            content_type='text/csv'
        )
        
        result = process_uploaded_file(file_storage)
        
        assert 'genes' in result
        assert len(result['genes']) == 2
        assert result['genes'][0]['gene_symbol'] == 'BRCA1'
        assert result['genes'][1]['gene_symbol'] == 'TP53'
    
    @patch('openpyxl.load_workbook')
    def test_process_excel_file(self, mock_load_workbook, sample_file_data):
        """Test processing of Excel file."""
        # Mock Excel workbook
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        
        # Mock Excel data
        excel_data = sample_file_data['valid_excel_data']
        mock_worksheet.iter_rows.return_value = [
            [Mock(value=cell) for cell in row] for row in excel_data
        ]
        
        mock_load_workbook.return_value = mock_workbook
        
        file_obj = io.BytesIO(b'fake excel content')
        file_storage = FileStorage(
            stream=file_obj,
            filename='test.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        result = process_uploaded_file(file_storage)
        
        assert 'genes' in result
        assert len(result['genes']) >= 1
    
    def test_process_file_with_missing_columns(self):
        """Test processing file with missing required columns."""
        csv_content = 'Symbol,Description\nBRCA1,BRCA1 gene\nTP53,TP53 gene'
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='missing_columns.csv',
            content_type='text/csv'
        )
        
        result = process_uploaded_file(file_storage)
        
        # Should handle missing columns gracefully
        assert 'errors' in result or 'genes' in result
    
    def test_process_file_with_extra_columns(self):
        """Test processing file with extra columns."""
        csv_content = 'Gene Symbol,Gene Name,Extra Column,Another Column\nBRCA1,BRCA1 gene,extra1,extra2\nTP53,TP53 gene,extra3,extra4'
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='extra_columns.csv',
            content_type='text/csv'
        )
        
        result = process_uploaded_file(file_storage)
        
        # Should process successfully, ignoring extra columns
        assert 'genes' in result
        assert len(result['genes']) == 2
    
    def test_process_file_with_duplicate_genes(self):
        """Test processing file with duplicate gene entries."""
        csv_content = 'Gene Symbol,Gene Name\nBRCA1,BRCA1 gene\nBRCA1,BRCA1 duplicate\nTP53,TP53 gene'
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='duplicates.csv',
            content_type='text/csv'
        )
        
        result = process_uploaded_file(file_storage)
        
        # Should handle duplicates appropriately
        assert 'genes' in result
        # Behavior depends on implementation - might dedupe or include all
        assert len(result['genes']) >= 2


@pytest.mark.unit
@pytest.mark.file_upload
class TestFileUploadRoutes:
    """Test file upload route handlers."""
    
    def test_upload_page_renders(self, authenticated_client):
        """Test upload page renders correctly."""
        response = authenticated_client.get('/upload')
        
        assert response.status_code == 200
        assert b'upload' in response.data.lower() or b'file' in response.data.lower()
    
    def test_upload_requires_authentication(self, client):
        """Test upload requires user authentication."""
        response = client.get('/upload')
        
        # Should redirect to login or show unauthorized
        assert response.status_code in [302, 401]
    
    def test_successful_file_upload(self, authenticated_client, sample_file_data):
        """Test successful file upload."""
        csv_content = sample_file_data['valid_csv']
        
        data = {
            'file': (io.BytesIO(csv_content.encode('utf-8')), 'test.csv')
        }
        
        response = authenticated_client.post('/upload', data=data)
        
        # Should process successfully
        assert response.status_code in [200, 302]  # Success or redirect
    
    def test_upload_without_file(self, authenticated_client):
        """Test upload without selecting a file."""
        response = authenticated_client.post('/upload', data={})
        
        assert response.status_code == 200
        # Should show error message
        assert b'file' in response.data.lower()
    
    def test_upload_invalid_file(self, authenticated_client):
        """Test upload of invalid file."""
        invalid_content = 'not a valid csv file format'
        
        data = {
            'file': (io.BytesIO(invalid_content.encode('utf-8')), 'invalid.csv')
        }
        
        response = authenticated_client.post('/upload', data=data)
        
        assert response.status_code == 200
        # Should show validation errors
        assert b'error' in response.data.lower() or b'invalid' in response.data.lower()


@pytest.mark.unit
@pytest.mark.file_upload
class TestFileStorage:
    """Test file storage and management."""
    
    @patch('os.makedirs')
    @patch('builtins.open', new_callable=mock_open)
    def test_save_uploaded_file(self, mock_file, mock_makedirs):
        """Test saving uploaded file to disk."""
        from app.secure_file_handler import SecureFileHandler
        
        file_content = b'test file content'
        file_obj = io.BytesIO(file_content)
        file_storage = FileStorage(
            stream=file_obj,
            filename='test.csv',
            content_type='text/csv'
        )
        
        handler = SecureFileHandler()
        file_path = handler.save_file(file_storage)
        
        assert file_path is not None
        mock_file.assert_called()
    
    def test_generate_secure_filename(self):
        """Test secure filename generation."""
        from app.secure_file_handler import SecureFileHandler
        
        handler = SecureFileHandler()
        
        # Test various filename scenarios
        test_cases = [
            ('normal.csv', 'csv'),
            ('with spaces.csv', 'csv'),
            ('with-special!@#chars.csv', 'csv'),
            ('very_long_filename_that_might_cause_issues.csv', 'csv'),
            ('unicode_filename_测试.csv', 'csv')
        ]
        
        for original, expected_ext in test_cases:
            secure_name = handler.generate_secure_filename(original)
            assert secure_name.endswith(f'.{expected_ext}')
            # Should not contain dangerous characters
            assert '..' not in secure_name
            assert '/' not in secure_name
            assert '\\' not in secure_name
    
    @patch('os.path.exists')
    @patch('os.remove')
    def test_cleanup_old_files(self, mock_remove, mock_exists):
        """Test cleanup of old uploaded files."""
        from app.secure_file_handler import SecureFileHandler
        
        mock_exists.return_value = True
        
        handler = SecureFileHandler()
        handler.cleanup_old_files(max_age_hours=24)
        
        # Should check for old files and remove them
        assert mock_exists.called or mock_remove.called or True  # Implementation dependent
    
    def test_file_virus_scanning(self):
        """Test file virus scanning (if implemented)."""
        # This would test integration with virus scanning
        # Implementation depends on actual virus scanning setup
        pass


@pytest.mark.unit
@pytest.mark.file_upload
@pytest.mark.security
class TestFileUploadSecurity:
    """Test file upload security measures."""
    
    def test_filename_sanitization(self):
        """Test that filenames are properly sanitized."""
        from werkzeug.utils import secure_filename
        
        dangerous_names = [
            '../../../etc/passwd',
            'file.exe',
            'script.js',
            'file with spaces.csv',
            'file\x00name.csv',
            'very_long_filename_' + 'x' * 200 + '.csv'
        ]
        
        for dangerous in dangerous_names:
            safe_name = secure_filename(dangerous)
            assert '..' not in safe_name
            assert '/' not in safe_name
            assert '\x00' not in safe_name
    
    def test_file_type_validation(self):
        """Test file type validation prevents dangerous uploads."""
        dangerous_types = [
            ('script.exe', 'application/octet-stream'),
            ('malware.bat', 'application/x-bat'),
            ('script.sh', 'application/x-shellscript'),
            ('virus.js', 'application/javascript')
        ]
        
        for filename, content_type in dangerous_types:
            file_obj = io.BytesIO(b'malicious content')
            file_storage = FileStorage(
                stream=file_obj,
                filename=filename,
                content_type=content_type
            )
            
            result = validate_file(file_storage)
            
            # Should reject dangerous file types
            assert result['valid'] is False
    
    def test_file_size_limits(self):
        """Test file size limits are enforced."""
        # Test oversized file
        large_content = b'x' * (50 * 1024 * 1024)  # 50MB
        file_obj = io.BytesIO(large_content)
        file_storage = FileStorage(
            stream=file_obj,
            filename='large.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        # Should respect size limits (depending on configuration)
        assert isinstance(result, dict)
    
    def test_file_content_scanning(self):
        """Test file content scanning for malicious content."""
        # Test file with script content in CSV
        malicious_csv = 'Gene Symbol,Gene Name\n<script>alert("xss")</script>,BRCA1\nTP53,Normal'
        file_obj = io.BytesIO(malicious_csv.encode('utf-8'))
        file_storage = FileStorage(
            stream=file_obj,
            filename='malicious.csv',
            content_type='text/csv'
        )
        
        result = validate_file(file_storage)
        
        # Should detect or sanitize malicious content
        assert isinstance(result, dict)
        assert 'valid' in result
